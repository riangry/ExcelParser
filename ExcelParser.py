from openpyxl import load_workbook


def estimate_parser(file):
    """
    Принимает файл excel
    парсит данные о сметах
    добавляет все элементы документа в список
    :param file:
    :return: list
    """
    wb = load_workbook(file)  # Открываем файл таблицы
    sheet = wb.active  # выбираем активный лист
    data = []  # создаем список для данных сметы
    results = ['Материалы', 'Машины и механизмы', 'ФОТ']  # создаем список для исключений
    row_count = sheet.max_row  # находим количество строк
    for row_number in range(1, row_count):  # итерация по каждой строке
        # находим значение каждого столбца в строке
        cell_a = str(sheet[f'A{row_number}'].value)
        cell_b = str(sheet[f'B{row_number}'].value)
        cell_c = str(sheet[f'C{row_number}'].value)
        cell_d = str(sheet[f'D{row_number}'].value)
        cell_f = str(sheet[f'F{row_number}'].value)
        cell_g = str(sheet[f'G{row_number}'].value)
        if "Раздел" in cell_a:  # Если слово Раздел в ячейке А
            data.append(cell_a.strip())  # добавляем в список
        elif cell_b == 'None':  # если ячейка Б пустая
            if cell_a != 'None' and cell_a.strip() not in results and \
                    "итог" not in cell_a.lower() and "ВСЕГО" not in cell_a:  # и ячейка А не пуста и значение А не в списке исключений
                data.append(cell_a.strip())  # добавляем в список
        if 'СРК' in cell_b:  # Если СРК в ячейке Б
            data.append([cell_b, cell_c])  # добавляем в список
        if "Цена по прайсу" in cell_b:  # Если Цена по прайсу в ячейке С
            data.append([cell_b, cell_c, cell_d, cell_f, cell_g])  # добавляем в список
        if "Механизмы" in cell_b:  # Если Механизмы в ячейке Б
            data.append([cell_b, cell_c, cell_d, cell_f, cell_g])  # добавляем в список
    return data  # Возвращаем список


section_ids = []
stage_ids = []
srk_ids = []
material_ids = []
mechanic_ids = []
data = estimate_parser('2.xlsx')
for item in data:
    print(item)
    if 'Раздел' in item:
        section_ids.append(data[data.index(item)])
    if isinstance(item, str) and 'Раздел' not in item:
        stage_ids.append(data[data.index(item)])
    if len(item) == 2:
        srk_ids.append(data[data.index(item)])
    if len(item) == 5 and 'Цена по прайсу' in item:
        material_ids.append(data[data.index(item)])
    if len(item) == 5 and 'Механизмы' in item:
        mechanic_ids.append(data[data.index(item)])
